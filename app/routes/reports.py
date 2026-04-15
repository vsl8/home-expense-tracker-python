"""
Reports API routes for generating expense reports and exports.
"""

import io
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, send_file
from sqlalchemy import func
from dateutil.relativedelta import relativedelta
from app import db
from app.models import Expense, ExpenseType

reports_bp = Blueprint("reports", __name__)


def get_date_range(period):
    """Calculate date range based on period type."""
    today = datetime.utcnow().date()

    if period == "daily":
        start_date = today
        end_date = today
    elif period == "weekly":
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif period == "monthly":
        start_date = today.replace(day=1)
        next_month = start_date + relativedelta(months=1)
        end_date = next_month - timedelta(days=1)
    elif period == "quarterly":
        quarter = (today.month - 1) // 3
        start_date = datetime(today.year, quarter * 3 + 1, 1).date()
        end_date = start_date + relativedelta(months=3) - timedelta(days=1)
    elif period == "half-yearly":
        if today.month <= 6:
            start_date = datetime(today.year, 1, 1).date()
            end_date = datetime(today.year, 6, 30).date()
        else:
            start_date = datetime(today.year, 7, 1).date()
            end_date = datetime(today.year, 12, 31).date()
    elif period == "yearly":
        start_date = datetime(today.year, 1, 1).date()
        end_date = datetime(today.year, 12, 31).date()
    else:
        # Default to monthly
        start_date = today.replace(day=1)
        next_month = start_date + relativedelta(months=1)
        end_date = next_month - timedelta(days=1)

    return start_date, end_date


@reports_bp.route("/summary", methods=["GET"])
def get_report_summary():
    """Get expense report summary for a given period."""
    period = request.args.get("period", "monthly")
    start_date, end_date = get_date_range(period)

    # Allow custom date range
    custom_start = request.args.get("start_date")
    custom_end = request.args.get("end_date")

    if custom_start:
        start_date = datetime.strptime(custom_start, "%Y-%m-%d").date()
    if custom_end:
        end_date = datetime.strptime(custom_end, "%Y-%m-%d").date()

    # Get expenses for the period
    expenses = Expense.query.filter(Expense.date >= start_date, Expense.date <= end_date).all()

    # Calculate totals
    total_amount = sum(e.amount for e in expenses)

    # Group by category
    category_totals = (
        db.session.query(
            ExpenseType.id,
            ExpenseType.name,
            ExpenseType.color,
            func.sum(Expense.amount).label("total"),
            func.count(Expense.id).label("count"),
        )
        .join(Expense)
        .filter(Expense.date >= start_date, Expense.date <= end_date)
        .group_by(ExpenseType.id)
        .all()
    )

    # Group by date for trend chart
    daily_totals = (
        db.session.query(Expense.date, func.sum(Expense.amount).label("total"))
        .filter(Expense.date >= start_date, Expense.date <= end_date)
        .group_by(Expense.date)
        .order_by(Expense.date)
        .all()
    )

    return jsonify(
        {
            "period": period,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "total_amount": round(total_amount, 2),
            "expense_count": len(expenses),
            "average_expense": round(total_amount / len(expenses), 2) if expenses else 0,
            "category_breakdown": [
                {
                    "id": cat[0],
                    "name": cat[1],
                    "color": cat[2],
                    "total": round(cat[3], 2),
                    "count": cat[4],
                    "percentage": round((cat[3] / total_amount) * 100, 1) if total_amount > 0 else 0,
                }
                for cat in category_totals
            ],
            "daily_trend": [{"date": day[0].isoformat(), "total": round(day[1], 2)} for day in daily_totals],
            "expenses": [e.to_dict() for e in expenses],
        }
    )


@reports_bp.route("/export/excel", methods=["GET"])
def export_to_excel():
    """Export expense report to Excel spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    period = request.args.get("period", "monthly")
    start_date, end_date = get_date_range(period)

    # Allow custom date range
    custom_start = request.args.get("start_date")
    custom_end = request.args.get("end_date")

    if custom_start:
        start_date = datetime.strptime(custom_start, "%Y-%m-%d").date()
    if custom_end:
        end_date = datetime.strptime(custom_end, "%Y-%m-%d").date()

    # Get expenses
    expenses = (
        Expense.query.filter(Expense.date >= start_date, Expense.date <= end_date).order_by(Expense.date.desc()).all()
    )

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Expense Report: {start_date.isoformat()} to {end_date.isoformat()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Summary section
    total_amount = sum(e.amount for e in expenses)
    ws["A3"] = "Total Expenses:"
    ws["B3"] = f"${total_amount:,.2f}"
    ws["A4"] = "Number of Transactions:"
    ws["B4"] = len(expenses)

    # Headers
    headers = ["Date", "Category", "Description", "Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, expense in enumerate(expenses, 7):
        ws.cell(row=row, column=1, value=expense.date.isoformat()).border = thin_border
        ws.cell(row=row, column=2, value=expense.expense_type.name).border = thin_border
        ws.cell(row=row, column=3, value=expense.description or "").border = thin_border
        amount_cell = ws.cell(row=row, column=4, value=expense.amount)
        amount_cell.border = thin_border
        amount_cell.number_format = "$#,##0.00"

    # Adjust column widths
    column_widths = [12, 15, 40, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"expense_report_{start_date}_{end_date}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route("/export/pdf", methods=["GET"])
def export_to_pdf():
    """Export expense report to PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch

    period = request.args.get("period", "monthly")
    start_date, end_date = get_date_range(period)

    # Allow custom date range
    custom_start = request.args.get("start_date")
    custom_end = request.args.get("end_date")

    if custom_start:
        start_date = datetime.strptime(custom_start, "%Y-%m-%d").date()
    if custom_end:
        end_date = datetime.strptime(custom_end, "%Y-%m-%d").date()

    # Get expenses
    expenses = (
        Expense.query.filter(Expense.date >= start_date, Expense.date <= end_date).order_by(Expense.date.desc()).all()
    )

    # Create PDF
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Heading1"], fontSize=18, spaceAfter=20, alignment=1  # Center
    )

    elements = []

    # Title
    elements.append(Paragraph("Home Expense Report", title_style))
    elements.append(Paragraph(f"Period: {start_date.isoformat()} to {end_date.isoformat()}", styles["Normal"]))
    elements.append(Spacer(1, 0.3 * inch))

    # Summary
    total_amount = sum(e.amount for e in expenses)
    summary_data = [
        ["Summary", ""],
        ["Total Expenses:", f"${total_amount:,.2f}"],
        ["Number of Transactions:", str(len(expenses))],
        ["Average per Transaction:", f"${(total_amount / len(expenses)):,.2f}" if expenses else "$0.00"],
    ]

    summary_table = Table(summary_data, colWidths=[2.5 * inch, 2 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#E9ECF1")),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("SPAN", (0, 0), (1, 0)),
            ]
        )
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5 * inch))

    # Expense details table
    if expenses:
        elements.append(Paragraph("Expense Details", styles["Heading2"]))
        elements.append(Spacer(1, 0.2 * inch))

        data = [["Date", "Category", "Description", "Amount"]]
        for expense in expenses:
            data.append(
                [
                    expense.date.isoformat(),
                    expense.expense_type.name,
                    (expense.description or "")[:40],
                    f"${expense.amount:,.2f}",
                ]
            )

        table = Table(data, colWidths=[1.2 * inch, 1.5 * inch, 2.5 * inch, 1 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                ]
            )
        )
        elements.append(table)

    doc.build(elements)
    output.seek(0)

    filename = f"expense_report_{start_date}_{end_date}.pdf"
    return send_file(output, mimetype="application/pdf", as_attachment=True, download_name=filename)


@reports_bp.route("/chart-data", methods=["GET"])
def get_chart_data():
    """Get data formatted for charts."""
    period = request.args.get("period", "monthly")
    chart_type = request.args.get("chart_type", "pie")  # pie, bar, line

    start_date, end_date = get_date_range(period)

    if chart_type == "pie":
        # Category breakdown for pie chart
        data = (
            db.session.query(ExpenseType.name, ExpenseType.color, func.sum(Expense.amount).label("total"))
            .join(Expense)
            .filter(Expense.date >= start_date, Expense.date <= end_date)
            .group_by(ExpenseType.id)
            .all()
        )

        return jsonify(
            {"labels": [d[0] for d in data], "colors": [d[1] for d in data], "data": [round(d[2], 2) for d in data]}
        )

    elif chart_type == "bar":
        # Category comparison for bar chart
        data = (
            db.session.query(ExpenseType.name, ExpenseType.color, func.sum(Expense.amount).label("total"))
            .join(Expense)
            .filter(Expense.date >= start_date, Expense.date <= end_date)
            .group_by(ExpenseType.id)
            .order_by(func.sum(Expense.amount).desc())
            .all()
        )

        return jsonify(
            {"labels": [d[0] for d in data], "colors": [d[1] for d in data], "data": [round(d[2], 2) for d in data]}
        )

    else:  # line chart - daily trend
        data = (
            db.session.query(Expense.date, func.sum(Expense.amount).label("total"))
            .filter(Expense.date >= start_date, Expense.date <= end_date)
            .group_by(Expense.date)
            .order_by(Expense.date)
            .all()
        )

        return jsonify({"labels": [d[0].isoformat() for d in data], "data": [round(d[1], 2) for d in data]})
